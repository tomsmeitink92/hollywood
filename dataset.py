from openpyxl import load_workbook
import pandas as pd


def load_data():
    wb = load_workbook(filename="films.xlsx")
    drop = ("Oscars won", "No of Oscars won", "Guardian film page", [None])

    def load(name):
        data = wb[name].values
        data = pd.DataFrame(columns=next(data), data=data)
        if "Cat" not in data.columns:
            data["Cat"] = name.title()
        for column in drop:
            data.drop(columns=column, errors="ignore", inplace=True)
        return data

    return pd.concat(load(name) for name in wb.sheetnames)


if __name__ == "__main__":
    df = load_data()
